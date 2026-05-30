import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="RM Planning Dashboard", layout="wide", page_icon="📦")

st.markdown("""
<style>
    .main-header {font-size: 2rem; font-weight: 700; color: #1e3a5f; margin-bottom: 0;}
    .sub-header {font-size: 1rem; color: #6b7280; margin-bottom: 1.5rem;}
    .section-header {font-size: 1.1rem; font-weight: 600; color: #374151; border-bottom: 2px solid #3b82f6; padding-bottom: 0.3rem; margin: 1.5rem 0 1rem;}
</style>
""", unsafe_allow_html=True)

st.markdown('<p class="main-header">📦 RM Planning Dashboard</p>', unsafe_allow_html=True)
st.markdown('<p class="sub-header">Upload Master SKU file and RM Stock file to compute producible cartons</p>', unsafe_allow_html=True)


@st.cache_data
def load_master(file_bytes):
    xl = pd.read_excel(file_bytes, sheet_name=None, header=None)

    sku_raw = xl.get("SKU Basic Details")
    if sku_raw is None:
        st.error("Sheet 'SKU Basic Details' not found in master file.")
        return None, None
    sku_df = sku_raw.iloc[3:].copy()
    sku_df.columns = sku_raw.iloc[2].tolist()
    sku_df = sku_df.dropna(how="all").reset_index(drop=True)
    rename_map = {}
    for c in sku_df.columns:
        c_str = str(c)
        if "Material ID" in c_str:
            rename_map[c] = "SKU"
        elif "Material Description" in c_str:
            rename_map[c] = "Description"
        elif "Factory" in c_str:
            rename_map[c] = "Factory"
        elif "Target FG" in c_str or "MTS" in c_str.upper():
            rename_map[c] = "MTS_Target"
    sku_df = sku_df.rename(columns=rename_map)
    sku_df = sku_df[["SKU", "Description", "Factory", "MTS_Target"]].copy()
    sku_df["SKU"] = sku_df["SKU"].apply(lambda x: str(x).strip() if pd.notna(x) else "")
    sku_df["MTS_Target"] = pd.to_numeric(sku_df["MTS_Target"], errors="coerce").fillna(0)

    bom_raw = xl.get("BOM - RM Details")
    if bom_raw is None:
        st.error("Sheet 'BOM - RM Details' not found in master file.")
        return None, None
    bom_df = bom_raw.iloc[4:].copy()
    bom_df.columns = bom_raw.iloc[3].tolist()
    bom_df = bom_df.dropna(how="all").reset_index(drop=True)
    rename_bom = {}
    for c in bom_df.columns:
        c_str = str(c)
        if "Material ID" in c_str and "SKU" in c_str:
            rename_bom[c] = "SKU"
        elif c_str == "Raw Material ID":
            rename_bom[c] = "RM_ID"
        elif c_str == "Material Description":
            rename_bom[c] = "Board_Desc"
        elif "Width" in c_str:
            rename_bom[c] = "Width_mm"
        elif "Length" in c_str:
            rename_bom[c] = "Length_mm"
        elif c_str == "GSM":
            rename_bom[c] = "GSM"
        elif "No. of Ups" in c_str:
            rename_bom[c] = "Num_Ups"
        elif "Wastage" in c_str:
            rename_bom[c] = "Wastage_Pct"
    bom_df = bom_df.rename(columns=rename_bom)
    keep = ["SKU", "RM_ID", "Board_Desc", "Width_mm", "Length_mm", "GSM", "Num_Ups", "Wastage_Pct"]
    bom_df = bom_df[[c for c in keep if c in bom_df.columns]].copy()
    bom_df["SKU"] = bom_df["SKU"].apply(lambda x: str(x).strip() if pd.notna(x) else "")
    bom_df["RM_ID"] = bom_df["RM_ID"].apply(lambda x: str(x).strip() if pd.notna(x) else "")
    for col in ["Width_mm", "Length_mm", "GSM", "Num_Ups", "Wastage_Pct"]:
        if col in bom_df.columns:
            bom_df[col] = pd.to_numeric(bom_df[col], errors="coerce")

    return sku_df, bom_df


@st.cache_data
def load_rm_stock(file_bytes):
    xl = pd.read_excel(file_bytes, sheet_name="Ageing Report")
    rm_df = xl[["Material", "Material Description", "Total Stock(Batch-wise)"]].copy()
    rm_df.columns = ["RM_ID", "RM_Desc", "Total_KG"]
    rm_df["RM_ID"] = rm_df["RM_ID"].apply(lambda x: str(x).strip() if pd.notna(x) else "")
    rm_df["Total_KG"] = pd.to_numeric(rm_df["Total_KG"], errors="coerce").fillna(0)
    return rm_df.groupby("RM_ID", as_index=False).agg(
        RM_Desc=("RM_Desc", "first"),
        Total_KG=("Total_KG", "sum"),
    )


def compute_cartons(kg, num_ups, length_mm, width_mm, gsm, wastage_pct):
    try:
        denom = length_mm * width_mm * gsm
        if denom == 0:
            return 0
        wastage = wastage_pct / 100.0 if wastage_pct > 1 else wastage_pct
        sheets = kg * 1_000_000_000 / denom
        return max(0, sheets * num_ups * (1 - wastage))
    except Exception:
        return 0


def build_output(sku_df, bom_df, rm_df):
    merged = bom_df.merge(sku_df, on="SKU", how="left")
    merged = merged.merge(rm_df, on="RM_ID", how="left")

    rm_sku_counts = bom_df.groupby("RM_ID")["SKU"].nunique().reset_index()
    rm_sku_counts.columns = ["RM_ID", "Sharing_Count"]
    merged = merged.merge(rm_sku_counts, on="RM_ID", how="left")
    merged["KG_Allocated"] = merged["Total_KG"] / merged["Sharing_Count"]

    rm_per_sku = bom_df.groupby("SKU")["RM_ID"].count().reset_index()
    rm_per_sku.columns = ["SKU", "RM_Component_Count"]
    merged = merged.merge(rm_per_sku, on="SKU", how="left")

    merged["Cartons_This_RM"] = merged.apply(
        lambda r: compute_cartons(
            r["KG_Allocated"],
            r.get("Num_Ups", 0) or 0,
            r.get("Length_mm", 0) or 0,
            r.get("Width_mm", 0) or 0,
            r.get("GSM", 0) or 0,
            r.get("Wastage_Pct", 0) or 0,
        ),
        axis=1,
    )

    sku_min = merged.groupby("SKU")["Cartons_This_RM"].min().reset_index()
    sku_min.columns = ["SKU", "Bottleneck_Cartons"]
    merged = merged.merge(sku_min, on="SKU", how="left")

    merged["MTS_x2"] = merged["MTS_Target"] * 2

    # Check: row-level — cartons this RM < MTS x2
    merged["Check"] = merged.apply(
        lambda r: "⚠ Below Qty" if pd.notna(r["Cartons_This_RM"]) and r["Cartons_This_RM"] < r["MTS_x2"] else "",
        axis=1,
    )

    merged["Cartons_This_RM"]   = merged["Cartons_This_RM"].round(0)
    merged["Bottleneck_Cartons"] = merged["Bottleneck_Cartons"].round(0)
    merged["KG_Allocated"]       = merged["KG_Allocated"].round(3)

    return merged


def style_output_df(df):
    """Yellow for below qty rows."""
    def highlight_check(row):
        if row.get("Check") == "⚠ Below Qty":
            return ["background-color: #fef9c3"] * len(row)
        return [""] * len(row)
    return df.style.apply(highlight_check, axis=1)


def style_bottleneck_df(df):
    """Red if bottleneck < MTS Target, yellow if bottleneck < MTS x2 but >= MTS Target."""
    def highlight(row):
        try:
            bn = float(row.get("Bottleneck Cartons", 0) or 0)
            mts = float(row.get("MTS Target", 0) or 0)
            mts2 = float(row.get("MTS x 2", 0) or 0)
        except Exception:
            return [""] * len(row)
        if bn < mts:
            return ["background-color: #fee2e2"] * len(row)   # red
        if bn < mts2:
            return ["background-color: #fef9c3"] * len(row)   # yellow
        return [""] * len(row)
    return df.style.apply(highlight, axis=1)


# Merged cols (same value for all RM rows of a SKU → merge in Excel)
MERGE_COLS = {"SKU", "Description", "Factory", "RM_Component_Count",
              "Bottleneck_Cartons", "MTS_Target", "MTS_x2"}

DISPLAY_COLS = [
    ("SKU",                 "SKU"),
    ("Description",         "Description"),
    ("Factory",             "Factory"),
    ("RM_Component_Count",  "RM Description (# Components)"),
    ("RM_ID",               "Board ID"),
    ("Board_Desc",          "Board Desc"),
    ("Total_KG",            "Board Quantity (KG)"),
    ("KG_Allocated",        "Board Qty Allocated (KG)"),
    ("Cartons_This_RM",     "Cartons (This RM)"),
    ("Bottleneck_Cartons",  "Bottleneck Cartons"),
    ("MTS_Target",          "MTS Target"),
    ("MTS_x2",              "MTS x 2"),
    ("Check",               "Check"),
]

DISPLAY_RENAME = {k: v for k, v in DISPLAY_COLS}


def to_single_factory_excel(fdf):
    """Build a single-sheet Excel workbook for one factory and return bytes."""
    buf = BytesIO()
    wb  = openpyxl.Workbook()
    ws  = wb.active

    header_fill  = PatternFill("solid", fgColor="1e3a5f")
    header_font  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    warn_fill    = PatternFill("solid", fgColor="FEF3C7")   # yellow
    red_fill     = PatternFill("solid", fgColor="FEE2E2")   # red
    normal_font  = Font(name="Arial", size=10)
    bold_font    = Font(name="Arial", size=10, bold=True)
    thin_border  = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    # Header row
    for col_idx, (_, header) in enumerate(DISPLAY_COLS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border
    ws.row_dimensions[1].height = 36

    excel_row = 2
    for sku, grp in fdf.groupby("SKU", sort=False):
        grp      = grp.reset_index(drop=True)
        n_rows   = len(grp)
        start_row = excel_row

        # Determine SKU-level highlight:
        # red  → bottleneck < MTS Target
        # yellow → bottleneck < MTS x2 (but >= MTS Target)
        try:
            bn   = float(grp["Bottleneck_Cartons"].iloc[0] or 0)
            mts  = float(grp["MTS_Target"].iloc[0] or 0)
            mts2 = float(grp["MTS_x2"].iloc[0] or 0)
        except Exception:
            bn = mts = mts2 = 0

        sku_fill = None
        if bn < mts:
            sku_fill = red_fill
        elif bn < mts2:
            sku_fill = warn_fill

        for i, (_, row) in enumerate(grp.iterrows()):
            is_row_warn = row.get("Check") == "⚠ Below Qty"
            for col_idx, (col_key, _) in enumerate(DISPLAY_COLS, start=1):
                if col_key in MERGE_COLS and i > 0:
                    cell = ws.cell(row=excel_row, column=col_idx)
                else:
                    val  = row.get(col_key, "")
                    val  = "" if pd.isna(val) else val
                    cell = ws.cell(row=excel_row, column=col_idx, value=val)

                cell.border    = thin_border
                cell.alignment = Alignment(
                    vertical="center",
                    horizontal="center" if col_key in MERGE_COLS else "left"
                )

                # Font
                if col_key == "Check" and is_row_warn:
                    cell.font = bold_font
                else:
                    cell.font = normal_font

                # Fill priority:
                # merged cols → use SKU-level fill
                # non-merged  → yellow if row is below qty, else SKU-level fill
                if col_key in MERGE_COLS:
                    if sku_fill:
                        cell.fill = sku_fill
                else:
                    if is_row_warn:
                        cell.fill = warn_fill
                    elif sku_fill:
                        cell.fill = sku_fill

            excel_row += 1

        # Merge SKU-level columns
        if n_rows > 1:
            end_row = start_row + n_rows - 1
            for col_idx, (col_key, _) in enumerate(DISPLAY_COLS, start=1):
                if col_key in MERGE_COLS:
                    ws.merge_cells(
                        start_row=start_row, start_column=col_idx,
                        end_row=end_row,     end_column=col_idx
                    )
                    mc = ws.cell(row=start_row, column=col_idx)
                    mc.alignment = Alignment(vertical="center", wrap_text=True, horizontal="center")
                    mc.border    = thin_border
                    if sku_fill:
                        mc.fill = sku_fill

    col_widths = [18, 42, 10, 20, 22, 38, 18, 22, 18, 18, 14, 12, 16]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─── Sidebar ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 📂 Upload Files")
    master_file = st.file_uploader("1. SKU Master Excel", type=["xlsx"], key="master")
    rm_file     = st.file_uploader("2. RM Stock Excel",   type=["xlsx"], key="rm")
    st.markdown("---")
    st.markdown("**Formula:**")
    st.latex(r"\text{Cartons} = \frac{kg \times 10^9}{L \times W \times GSM} \times N_{ups} \times (1-\text{wastage\%})")
    st.markdown("---")
    st.markdown("**Sharing:** RM split equally among sharing SKUs.  \n**Bottleneck:** min cartons across all RM for a SKU.")
    st.markdown("---")
    st.markdown("**Excel highlights:**  \n🟡 Yellow = any RM row below MTS×2  \n🔴 Red = bottleneck < MTS Target")

# ─── Main ─────────────────────────────────────────────────────────────────────
if not master_file or not rm_file:
    st.info("👈 Upload both files in the sidebar to begin.")
    st.stop()

with st.spinner("Loading master file..."):
    sku_df, bom_df = load_master(master_file.read())
with st.spinner("Loading RM stock file..."):
    rm_df = load_rm_stock(rm_file.read())

if sku_df is None or bom_df is None:
    st.stop()

with st.spinner("Computing cartons..."):
    result = build_output(sku_df, bom_df, rm_df)

# ─── Metrics ──────────────────────────────────────────────────────────────────
total_skus = result["SKU"].nunique()
below_qty  = (result["Check"] == "⚠ Below Qty").sum()
factories  = result["Factory"].dropna().unique().tolist()
rm_shared  = (result.groupby("RM_ID")["SKU"].nunique() > 1).sum()

c1, c2, c3, c4 = st.columns(4)
c1.metric("Total SKUs",          total_skus)
c2.metric("RM Rows Below MTS×2", int(below_qty), delta_color="inverse")
c3.metric("Factories",           len(factories))
c4.metric("Shared RM Materials", int(rm_shared))
st.markdown("---")

# ─── Per-factory tabs ─────────────────────────────────────────────────────────
all_factories = sorted(result["Factory"].dropna().unique().tolist())
tabs = st.tabs([f"🏭 {f}" for f in all_factories] + ["📊 All SKUs"])

for tab, factory in zip(tabs[:-1], all_factories):
    with tab:
        fdf = result[result["Factory"] == factory].copy()
        display_df = fdf[list(DISPLAY_RENAME.keys())].rename(columns=DISPLAY_RENAME)

        warn_count = (fdf["Check"] == "⚠ Below Qty").sum()
        col_a, col_b, col_c = st.columns(3)
        col_a.metric("SKUs", fdf["SKU"].nunique())
        col_b.metric("RM Rows Below MTS×2", int(warn_count))
        col_c.metric("RM Components", len(fdf))

        st.markdown(f'<div class="section-header">SKU → RM Producibility Table — {factory}</div>',
                    unsafe_allow_html=True)
        st.dataframe(style_output_df(display_df), use_container_width=True, hide_index=True)

        # ── Bottleneck table ──
        st.markdown(f'<div class="section-header">📌 Bottleneck Summary — {factory}</div>',
                    unsafe_allow_html=True)

        # Build one row per SKU
        bottleneck_rows = []
        for sku, grp in fdf.groupby("SKU", sort=False):
            min_idx = grp["Cartons_This_RM"].idxmin()
            limiting_rm = grp.loc[min_idx, "Board_Desc"] if min_idx in grp.index else ""
            bottleneck_rows.append({
                "SKU":                sku,
                "Description":        grp["Description"].iloc[0],
                "Bottleneck Cartons": int(grp["Bottleneck_Cartons"].iloc[0]),
                "Limiting Board":     limiting_rm,
                "MTS Target":         int(grp["MTS_Target"].iloc[0]),
                "MTS x 2":            int(grp["MTS_x2"].iloc[0]),
                "Status": (
                    "🔴 Below MTS" if grp["Bottleneck_Cartons"].iloc[0] < grp["MTS_Target"].iloc[0]
                    else "🟡 Below MTS×2" if grp["Bottleneck_Cartons"].iloc[0] < grp["MTS_x2"].iloc[0]
                    else "✅ OK"
                ),
            })
        bn_df = pd.DataFrame(bottleneck_rows)
        st.dataframe(style_bottleneck_df(bn_df), use_container_width=True, hide_index=True)

with tabs[-1]:
    display_all = result[list(DISPLAY_RENAME.keys())].rename(columns=DISPLAY_RENAME)
    st.dataframe(style_output_df(display_all), use_container_width=True, hide_index=True)

# ─── Download ─────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown('<div class="section-header">⬇ Download Output</div>', unsafe_allow_html=True)

factory_dfs = {f: result[result["Factory"] == f].copy() for f in all_factories}

dl_cols = st.columns(len(all_factories))
for col, factory in zip(dl_cols, all_factories):
    fdf = factory_dfs.get(factory, pd.DataFrame(columns=result.columns))
    excel_bytes = to_single_factory_excel(fdf)
    col.download_button(
        f"📥 board_status_{factory.lower()}.xlsx",
        data=excel_bytes,
        file_name=f"board_status_{factory.lower()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

# ─── Previews ─────────────────────────────────────────────────────────────────
with st.expander("🔍 Preview: Loaded SKU Master"):
    st.dataframe(sku_df, use_container_width=True, hide_index=True)
with st.expander("🔍 Preview: Loaded BOM"):
    st.dataframe(bom_df, use_container_width=True, hide_index=True)
with st.expander("🔍 Preview: Loaded RM Stock"):
    st.dataframe(rm_df, use_container_width=True, hide_index=True)